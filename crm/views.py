"""SERVIAC CRM - Views"""
from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.http import JsonResponse
from django.utils import timezone
from django.db.models import Count, Q, Sum, F
from datetime import timedelta

from .models import Customer, Product, Category, OrderInbox, Notification, OrderInboxStatus, CustomerType
from .forms import CustomerForm, ProductForm


class DashboardView(LoginRequiredMixin, TemplateView):
    """Tableau de bord CRM principal"""
    template_name = 'crm/dashboard.html'
    
    def get_context_data(self, **kwargs):
        from .models import Order, Invoice, OrderStatus, InvoiceStatus
        
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Tableau de bord'
        
        # Stats de base
        context['new_orders_count'] = OrderInbox.objects.filter(status=OrderInboxStatus.NEW).count()
        context['customers_count'] = Customer.objects.filter(is_active=True).count()
        
        # Commandes en cours (non facturées, non annulées)
        context['orders_in_progress'] = Order.objects.filter(
            status__in=[OrderStatus.DRAFT, OrderStatus.CONFIRMED, OrderStatus.PROCESSING, 
                       OrderStatus.READY, OrderStatus.DELIVERED]
        ).count()
        
        # Factures impayées
        context['unpaid_invoices'] = Invoice.objects.filter(
            status__in=[InvoiceStatus.DRAFT, InvoiceStatus.SENT, InvoiceStatus.PARTIAL, InvoiceStatus.OVERDUE]
        ).count()
        
        # Dernières commandes
        context['recent_orders'] = Order.objects.select_related('customer').order_by('-created_at')[:5]
        
        # Notifications
        context['recent_notifications'] = Notification.objects.filter(
            user=self.request.user, is_read=False
        )[:5]
        
        # Stats mensuelles
        today = timezone.now()
        month_start = today.replace(day=1, hour=0, minute=0, second=0)
        
        monthly_invoices = Invoice.objects.filter(
            created_at__gte=month_start,
            status__in=[InvoiceStatus.SENT, InvoiceStatus.PAID, InvoiceStatus.PARTIAL]
        )
        context['monthly_revenue'] = monthly_invoices.aggregate(total=Sum('total'))['total'] or 0
        context['monthly_orders'] = Order.objects.filter(created_at__gte=month_start).count()
        
        # Total créances clients
        context['total_receivables'] = Customer.objects.filter(
            is_active=True, balance__gt=0
        ).aggregate(total=Sum('balance'))['total'] or 0
        
        return context


# --- Customers ---
class CustomerListView(LoginRequiredMixin, ListView):
    model = Customer
    template_name = 'crm/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 20
    
    def get_queryset(self):
        qs = Customer.objects.all()
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(Q(name__icontains=search) | Q(phone__icontains=search))
        return qs


class CustomerCreateView(LoginRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'crm/customer_form.html'
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Client créé avec succès')
        return super().form_valid(form)
    
    def get_success_url(self):
        return self.object.get_absolute_url()


class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'crm/customer_detail.html'
    context_object_name = 'customer'


class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    model = Customer
    form_class = CustomerForm
    template_name = 'crm/customer_form.html'
    
    def form_valid(self, form):
        messages.success(self.request, 'Client modifié avec succès')
        return super().form_valid(form)
    
    def get_success_url(self):
        return self.object.get_absolute_url()


# --- Products ---
class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'crm/product_list.html'
    context_object_name = 'products'
    paginate_by = 20


class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'crm/product_form.html'
    
    def form_valid(self, form):
        messages.success(self.request, 'Produit créé avec succès')
        return super().form_valid(form)
    
    def get_success_url(self):
        return self.object.get_absolute_url()


class ProductDetailView(LoginRequiredMixin, DetailView):
    model = Product
    template_name = 'crm/product_detail.html'
    context_object_name = 'product'


class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'crm/product_form.html'
    
    def form_valid(self, form):
        messages.success(self.request, 'Produit modifié avec succès')
        return super().form_valid(form)
    
    def get_success_url(self):
        return self.object.get_absolute_url()


# --- Order Inbox ---
class OrderInboxListView(LoginRequiredMixin, ListView):
    model = OrderInbox
    template_name = 'crm/inbox_list.html'
    context_object_name = 'orders'
    paginate_by = 20
    
    def get_queryset(self):
        qs = OrderInbox.objects.select_related('customer', 'assigned_to')
        status = self.request.GET.get('status', 'all')
        if status != 'all':
            qs = qs.filter(status=status)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status'] = self.request.GET.get('status', 'all')
        context['new_count'] = OrderInbox.objects.filter(status=OrderInboxStatus.NEW).count()
        return context


class OrderInboxDetailView(LoginRequiredMixin, DetailView):
    model = OrderInbox
    template_name = 'crm/inbox_detail.html'
    context_object_name = 'order'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Récupérer les produits pour afficher les détails
        if self.object.items:
            product_ids = [item.get('product_id') for item in self.object.items if item.get('product_id')]
            context['products'] = {p.id: p for p in Product.objects.filter(id__in=product_ids)}
        return context


class OrderInboxValidateView(LoginRequiredMixin, View):
    def post(self, request, pk):
        order = get_object_or_404(OrderInbox, pk=pk)
        order.status = OrderInboxStatus.VALIDATED
        order.processed_at = timezone.now()
        order.processed_by = request.user
        order.save()
        messages.success(request, f'Commande #{pk} validée avec succès')
        return redirect('crm:inbox_list')


class OrderInboxRejectView(LoginRequiredMixin, View):
    def post(self, request, pk):
        order = get_object_or_404(OrderInbox, pk=pk)
        order.status = OrderInboxStatus.REJECTED
        order.rejection_reason = request.POST.get('reason', '')
        order.processed_at = timezone.now()
        order.processed_by = request.user
        order.save()
        messages.warning(request, f'Commande #{pk} rejetée')
        return redirect('crm:inbox_list')


# --- Notifications ---
class NotificationListView(LoginRequiredMixin, ListView):
    model = Notification
    template_name = 'crm/notification_list.html'
    context_object_name = 'notifications'
    paginate_by = 20
    
    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user)


@login_required
def mark_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'success': True})
    return JsonResponse({'success': False}, status=405)


# ============================================================
# PHASE 2: GESTION COMMERCIALE (Orders & Invoices)
# ============================================================

from .models import Order, OrderItem, Invoice, InvoiceItem, OrderStatus, PaymentStatus, InvoiceStatus
from .forms import OrderForm, OrderItemFormSet


class OrderInboxConvertView(LoginRequiredMixin, View):
    """Convertit une demande inbox en commande"""
    
    def get(self, request, pk):
        """Redirige vers la page de détail si accès en GET"""
        return redirect('crm:inbox_detail', pk=pk)
    
    def post(self, request, pk):
        inbox = get_object_or_404(OrderInbox, pk=pk)
        
        # Créer ou récupérer le client
        if inbox.customer:
            customer = inbox.customer
        else:
            # Créer un nouveau client depuis le prospect
            customer = Customer.objects.create(
                name=inbox.prospect_name,
                phone=inbox.prospect_phone,
                email=inbox.prospect_email or '',
                address=inbox.prospect_address,
                customer_type=inbox.prospect_type,
                created_by=request.user
            )
            inbox.customer = customer
        
        # Créer la commande
        order = Order.objects.create(
            customer=customer,
            delivery_address=inbox.delivery_address or customer.address,
            source_inbox=inbox,
            customer_notes=inbox.notes,
            created_by=request.user
        )
        
        # Créer les lignes de commande
        for item in inbox.items:
            try:
                product = Product.objects.get(pk=item.get('product_id'))
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=item.get('quantity', 1),
                    unit_price=item.get('unit_price', product.get_price_for_customer(customer))
                )
            except Product.DoesNotExist:
                continue
        
        # Recalculer les totaux
        order.calculate_totals()
        order.save()
        
        # Mettre à jour l'inbox
        inbox.status = OrderInboxStatus.CONVERTED
        inbox.processed_at = timezone.now()
        inbox.processed_by = request.user
        inbox.save()
        
        messages.success(request, f'Commande {order.number} créée avec succès')
        return redirect('crm:order_detail', pk=order.pk)


class OrderListView(LoginRequiredMixin, ListView):
    model = Order
    template_name = 'crm/order_list.html'
    context_object_name = 'orders'
    paginate_by = 20
    
    def get_queryset(self):
        qs = Order.objects.select_related('customer')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = OrderStatus.choices
        context['current_status'] = self.request.GET.get('status', '')
        return context


class OrderCreateView(LoginRequiredMixin, CreateView):
    model = Order
    form_class = OrderForm
    template_name = 'crm/order_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.filter(is_active=True)
        context['is_new'] = True
        return context
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        self.object = form.save()
        
        # Traiter les lignes de commande depuis le formulaire
        items_data = self.request.POST.getlist('item_product')
        quantities = self.request.POST.getlist('item_quantity')
        prices = self.request.POST.getlist('item_price')
        
        for i, product_id in enumerate(items_data):
            if product_id:
                try:
                    product = Product.objects.get(pk=product_id)
                    OrderItem.objects.create(
                        order=self.object,
                        product=product,
                        quantity=quantities[i] if i < len(quantities) else 1,
                        unit_price=prices[i] if i < len(prices) else product.price_b2c
                    )
                except (Product.DoesNotExist, ValueError):
                    continue
        
        self.object.calculate_totals()
        self.object.save()
        
        messages.success(self.request, f'Commande {self.object.number} créée')
        return redirect(self.object.get_absolute_url())


class OrderDetailView(LoginRequiredMixin, DetailView):
    model = Order
    template_name = 'crm/order_detail.html'
    context_object_name = 'order'


class OrderUpdateView(LoginRequiredMixin, UpdateView):
    model = Order
    form_class = OrderForm
    template_name = 'crm/order_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.filter(is_active=True)
        context['is_new'] = False
        return context


class OrderConfirmView(LoginRequiredMixin, View):
    """Confirme une commande"""
    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        if order.status == OrderStatus.DRAFT:
            order.status = OrderStatus.CONFIRMED
            order.save()
            
            # Mettre à jour le solde client
            order.customer.balance += order.total
            order.customer.save()
            
            messages.success(request, f'Commande {order.number} confirmée')
        return redirect('crm:order_detail', pk=pk)


class OrderInvoiceView(LoginRequiredMixin, View):
    """Crée une facture depuis une commande"""
    def post(self, request, pk):
        order = get_object_or_404(Order, pk=pk)
        if order.status in [OrderStatus.CONFIRMED, OrderStatus.DELIVERED]:
            invoice = Invoice.create_from_order(order, user=request.user)
            order.status = OrderStatus.INVOICED
            order.save()
            messages.success(request, f'Facture {invoice.number} créée')
            return redirect('crm:invoice_detail', pk=invoice.pk)
        messages.error(request, 'Impossible de facturer cette commande')
        return redirect('crm:order_detail', pk=pk)


class InvoiceListView(LoginRequiredMixin, ListView):
    model = Invoice
    template_name = 'crm/invoice_list.html'
    context_object_name = 'invoices'
    paginate_by = 20
    
    def get_queryset(self):
        qs = Invoice.objects.select_related('customer', 'order')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs


class InvoiceDetailView(LoginRequiredMixin, DetailView):
    model = Invoice
    template_name = 'crm/invoice_detail.html'
    context_object_name = 'invoice'


class InvoicePDFView(LoginRequiredMixin, View):
    """Génère le PDF d'une facture"""
    def get(self, request, pk):
        from django.http import HttpResponse
        from django.template.loader import render_to_string
        
        invoice = get_object_or_404(Invoice, pk=pk)
        
        # Générer HTML
        html = render_to_string('crm/invoice_pdf.html', {
            'invoice': invoice,
            'company': {
                'name': 'SERVIAC GROUP SUARL',
                'address': 'Abidjan, Côte d\'Ivoire',
                'phone': '+225 XX XX XX XX XX',
                'email': 'contact@serviac-group.com',
            }
        })
        
        # Pour l'instant, retourner HTML (WeasyPrint pour PDF réel)
        response = HttpResponse(html, content_type='text/html')
        return response


class InvoicePaymentView(LoginRequiredMixin, View):
    """Enregistre un paiement sur une facture"""
    def post(self, request, pk):
        invoice = get_object_or_404(Invoice, pk=pk)
        amount = request.POST.get('amount')
        method = request.POST.get('method', 'cash')
        transaction_ref = request.POST.get('transaction_ref', '')
        
        try:
            amount = Decimal(amount)
            if amount <= 0:
                raise ValueError("Montant invalide")
                
            # Trouver la caisse ouverte
            cash_register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
            
            Payment.objects.create(
                customer=invoice.customer,
                invoice=invoice,
                amount=amount,
                payment_method=method,
                transaction_ref=transaction_ref,
                cash_register=cash_register,
                created_by=request.user
            )
            
            if cash_register:
                cash_register.recalculate_totals()
            
            messages.success(request, f'Paiement de {amount} FCFA enregistré')
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f'Erreur: {str(e)}')
        
        return redirect('crm:invoice_detail', pk=pk)


# ============================================================
# PHASE 3: GESTION FINANCIÈRE
# ============================================================

from .models import (Payment, PaymentMethod, PaymentType, CustomerLedger, 
                     CustomerCredit, PaymentSchedule, CashRegister, CashRegisterStatus, CashMovement)
from decimal import Decimal, InvalidOperation


class PaymentListView(LoginRequiredMixin, ListView):
    model = Payment
    template_name = 'crm/payment_list.html'
    context_object_name = 'payments'
    paginate_by = 30
    
    def get_queryset(self):
        qs = Payment.objects.select_related('customer', 'invoice')
        method = self.request.GET.get('method')
        if method:
            qs = qs.filter(payment_method=method)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['payment_methods'] = PaymentMethod.choices
        return context


class PaymentCreateView(LoginRequiredMixin, CreateView):
    model = Payment
    template_name = 'crm/payment_form.html'
    fields = ['customer', 'invoice', 'amount', 'payment_method', 'transaction_ref', 'notes']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['customers'] = Customer.objects.filter(is_active=True, balance__gt=0)
        context['invoices'] = Invoice.objects.filter(status__in=[InvoiceStatus.SENT, InvoiceStatus.PARTIAL])
        context['payment_methods'] = PaymentMethod.choices
        return context
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        # Associer à la caisse ouverte
        cash_register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
        if cash_register:
            form.instance.cash_register = cash_register
        
        response = super().form_valid(form)
        
        if cash_register:
            cash_register.recalculate_totals()
        
        messages.success(self.request, f'Paiement {self.object.reference} enregistré')
        return response
    
    def get_success_url(self):
        return reverse('crm:payment_detail', kwargs={'pk': self.object.pk})


class PaymentDetailView(LoginRequiredMixin, DetailView):
    model = Payment
    template_name = 'crm/payment_detail.html'
    context_object_name = 'payment'


class CustomerLedgerView(LoginRequiredMixin, ListView):
    """Grand livre d'un client"""
    model = CustomerLedger
    template_name = 'crm/customer_ledger.html'
    context_object_name = 'entries'
    paginate_by = 50
    
    def get_queryset(self):
        self.customer = get_object_or_404(Customer, pk=self.kwargs['pk'])
        return CustomerLedger.objects.filter(customer=self.customer)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['customer'] = self.customer
        return context


class CashRegisterView(LoginRequiredMixin, TemplateView):
    """Vue principale de la caisse"""
    template_name = 'crm/cash_register.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Caisse ouverte ?
        context['current_register'] = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
        
        if context['current_register']:
            # Récupérer les paiements du jour
            context['today_payments'] = Payment.objects.filter(
                cash_register=context['current_register']
            ).select_related('customer', 'invoice')
            
            # Mouvements de caisse
            context['movements'] = context['current_register'].movements.all()
        
        # Clients avec solde pour encaissement rapide
        context['customers_with_balance'] = Customer.objects.filter(
            is_active=True, balance__gt=0
        ).order_by('-balance')[:10]
        
        context['payment_methods'] = PaymentMethod.choices
        
        return context


class CashRegisterOpenView(LoginRequiredMixin, View):
    """Ouvrir une nouvelle session de caisse"""
    
    def get(self, request):
        """Redirige vers la page caisse si accès en GET"""
        return redirect('crm:cash_register')
    
    def post(self, request):
        # Vérifier qu'aucune caisse n'est ouverte
        if CashRegister.objects.filter(status=CashRegisterStatus.OPEN).exists():
            messages.error(request, 'Une caisse est déjà ouverte')
            return redirect('crm:cash_register')
        
        opening_balance = request.POST.get('opening_balance', 0)
        try:
            opening_balance = Decimal(opening_balance)
        except:
            opening_balance = Decimal('0')
        
        CashRegister.objects.create(
            opening_balance=opening_balance,
            expected_balance=opening_balance,
            opened_by=request.user
        )
        
        messages.success(request, 'Caisse ouverte avec succès')
        return redirect('crm:cash_register')


class CashRegisterCloseView(LoginRequiredMixin, View):
    """Clôturer la session de caisse"""
    
    def get(self, request):
        """Redirige vers la page caisse si accès en GET"""
        return redirect('crm:cash_register')
    
    def post(self, request):
        register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
        if not register:
            messages.error(request, 'Aucune caisse ouverte')
            return redirect('crm:cash_register')
        
        closing_balance = request.POST.get('closing_balance', 0)
        try:
            closing_balance = Decimal(closing_balance)
        except:
            closing_balance = register.expected_balance
        
        register.close(closing_balance, request.user)
        
        diff = register.difference
        if diff and diff != 0:
            if diff > 0:
                messages.warning(request, f'Caisse clôturée avec un excédent de {diff} FCFA')
            else:
                messages.warning(request, f'Caisse clôturée avec un déficit de {abs(diff)} FCFA')
        else:
            messages.success(request, 'Caisse clôturée avec succès')
        
        return redirect('crm:cash_register')


class CashRegisterPaymentView(LoginRequiredMixin, View):
    """Encaisser rapidement depuis la caisse"""
    def post(self, request):
        register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
        if not register:
            messages.error(request, 'Ouvrez d\'abord la caisse')
            return redirect('crm:cash_register')
        
        customer_id = request.POST.get('customer')
        amount = request.POST.get('amount')
        method = request.POST.get('method', 'cash')
        invoice_id = request.POST.get('invoice')
        transaction_ref = request.POST.get('transaction_ref', '')
        
        try:
            customer = Customer.objects.get(pk=customer_id)
            amount = Decimal(amount)
            invoice = Invoice.objects.get(pk=invoice_id) if invoice_id else None
            
            Payment.objects.create(
                customer=customer,
                invoice=invoice,
                amount=amount,
                payment_method=method,
                transaction_ref=transaction_ref,
                cash_register=register,
                created_by=request.user
            )
            
            register.recalculate_totals()
            messages.success(request, f'Encaissement de {amount} FCFA effectué')
            
        except (Customer.DoesNotExist, ValueError, InvalidOperation) as e:
            messages.error(request, f'Erreur: {str(e)}')
        
        return redirect('crm:cash_register')


class CashMovementView(LoginRequiredMixin, View):
    """Enregistrer un mouvement de caisse (entrée/sortie hors vente)"""
    def post(self, request):
        register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
        if not register:
            messages.error(request, 'Ouvrez d\'abord la caisse')
            return redirect('crm:cash_register')
        
        movement_type = request.POST.get('movement_type')
        amount = request.POST.get('amount')
        reason = request.POST.get('reason', '')
        
        try:
            amount = Decimal(amount)
            CashMovement.objects.create(
                cash_register=register,
                movement_type=movement_type,
                amount=amount,
                reason=reason,
                created_by=request.user
            )
            
            # Mettre à jour le solde théorique
            if movement_type == 'in':
                register.expected_balance += amount
            else:
                register.expected_balance -= amount
            register.save()
            
            messages.success(request, f'Mouvement enregistré: {amount} FCFA')
        except (ValueError, InvalidOperation) as e:
            messages.error(request, f'Erreur: {str(e)}')
        
        return redirect('crm:cash_register')


class CashRegisterHistoryView(LoginRequiredMixin, ListView):
    """Historique des sessions de caisse"""
    model = CashRegister
    template_name = 'crm/cash_register_history.html'
    context_object_name = 'registers'
    paginate_by = 20


class AgedBalanceView(LoginRequiredMixin, TemplateView):
    """Balance âgée des clients"""
    template_name = 'crm/aged_balance.html'
    
    def get_context_data(self, **kwargs):
        from datetime import timedelta
        context = super().get_context_data(**kwargs)
        
        today = timezone.now().date()
        
        # Récupérer toutes les factures non payées
        unpaid_invoices = Invoice.objects.filter(
            status__in=[InvoiceStatus.SENT, InvoiceStatus.PARTIAL, InvoiceStatus.OVERDUE]
        ).select_related('customer')
        
        # Calculer les tranches d'âge
        aged_data = []
        totals = {'current': 0, 'days_30': 0, 'days_60': 0, 'days_90': 0, 'over_90': 0, 'total': 0}
        
        customers_balance = {}
        for inv in unpaid_invoices:
            remaining = inv.total - inv.amount_paid
            if remaining <= 0:
                continue
                
            days_old = (today - inv.invoice_date).days
            
            if inv.customer_id not in customers_balance:
                customers_balance[inv.customer_id] = {
                    'customer': inv.customer,
                    'current': Decimal('0'),
                    'days_30': Decimal('0'),
                    'days_60': Decimal('0'),
                    'days_90': Decimal('0'),
                    'over_90': Decimal('0'),
                    'total': Decimal('0')
                }
            
            cb = customers_balance[inv.customer_id]
            cb['total'] += remaining
            totals['total'] += remaining
            
            if days_old <= 30:
                cb['current'] += remaining
                totals['current'] += remaining
            elif days_old <= 60:
                cb['days_30'] += remaining
                totals['days_30'] += remaining
            elif days_old <= 90:
                cb['days_60'] += remaining
                totals['days_60'] += remaining
            elif days_old <= 120:
                cb['days_90'] += remaining
                totals['days_90'] += remaining
            else:
                cb['over_90'] += remaining
                totals['over_90'] += remaining
        
        # Trier par total décroissant
        aged_data = sorted(customers_balance.values(), key=lambda x: x['total'], reverse=True)
        
        context['aged_data'] = aged_data
        context['totals'] = totals
        
        return context


# ============================================================
# PHASE 4: GESTION STOCK INTELLIGENTE
# ============================================================

from .models import (Shipment, ShipmentItem, ShipmentStatus, StockMovement, StockMovementType,
                     CustomerScore, PreOrder, PreOrderStatus, StockReservation,
                     InventoryControl, InventoryLine, InventoryControlStatus)


class StockDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard stock avec vue d'ensemble"""
    template_name = 'crm/stock_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Produits avec stock
        products = Product.objects.filter(is_active=True)
        context['products'] = products
        
        # Alertes stock bas
        context['low_stock'] = products.filter(
            stock_quantity__lte=F('alert_threshold')
        )
        
        # Expéditions en transit
        context['shipments_in_transit'] = Shipment.objects.filter(
            status__in=[ShipmentStatus.SHIPPED, ShipmentStatus.IN_TRANSIT]
        )
        
        # Stock en transit
        transit_items = ShipmentItem.objects.filter(
            shipment__status__in=[ShipmentStatus.SHIPPED, ShipmentStatus.IN_TRANSIT]
        ).values('product__name', 'product__code').annotate(
            total=Sum('quantity')
        )
        context['stock_in_transit'] = transit_items
        
        # Pré-commandes en attente
        context['pending_preorders'] = PreOrder.objects.filter(status=PreOrderStatus.PENDING).count()
        
        # Réservations actives
        context['active_reservations'] = StockReservation.objects.filter(is_active=True).aggregate(
            total=Sum('quantity')
        )['total'] or 0
        
        # Totaux
        context['total_stock'] = products.aggregate(total=Sum('stock_quantity'))['total'] or 0
        
        return context


class StockMovementListView(LoginRequiredMixin, ListView):
    """Historique des mouvements de stock"""
    model = StockMovement
    template_name = 'crm/stock_movements.html'
    context_object_name = 'movements'
    paginate_by = 50
    
    def get_queryset(self):
        qs = StockMovement.objects.select_related('product', 'shipment', 'order')
        product = self.request.GET.get('product')
        if product:
            qs = qs.filter(product_id=product)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.filter(is_active=True)
        return context


class ShipmentListView(LoginRequiredMixin, ListView):
    """Liste des expéditions"""
    model = Shipment
    template_name = 'crm/shipment_list.html'
    context_object_name = 'shipments'
    paginate_by = 20
    
    def get_queryset(self):
        qs = Shipment.objects.prefetch_related('items__product')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = ShipmentStatus.choices
        return context


class ShipmentCreateView(LoginRequiredMixin, CreateView):
    """Créer une nouvelle expédition"""
    model = Shipment
    template_name = 'crm/shipment_form.html'
    fields = ['departure_date', 'estimated_arrival', 'transporter', 'vehicle_info', 'notes']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.filter(is_active=True)
        return context
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        self.object = form.save()
        
        # Traiter les lignes d'expédition
        products = self.request.POST.getlist('item_product')
        quantities = self.request.POST.getlist('item_quantity')
        
        for i, product_id in enumerate(products):
            if product_id and i < len(quantities):
                try:
                    product = Product.objects.get(pk=product_id)
                    qty = Decimal(quantities[i])
                    if qty > 0:
                        ShipmentItem.objects.create(
                            shipment=self.object,
                            product=product,
                            quantity=qty
                        )
                except (Product.DoesNotExist, ValueError):
                    continue
        
        messages.success(self.request, f'Expédition {self.object.reference} créée')
        return redirect('crm:shipment_detail', pk=self.object.pk)


class ShipmentDetailView(LoginRequiredMixin, DetailView):
    model = Shipment
    template_name = 'crm/shipment_detail.html'
    context_object_name = 'shipment'


class ShipmentShipView(LoginRequiredMixin, View):
    """Marquer une expédition comme expédiée"""
    def post(self, request, pk):
        shipment = get_object_or_404(Shipment, pk=pk)
        if shipment.status == ShipmentStatus.PREPARING:
            shipment.status = ShipmentStatus.SHIPPED
            shipment.departure_date = timezone.now().date()
            shipment.save()
            messages.success(request, f'Expédition {shipment.reference} marquée comme expédiée')
        return redirect('crm:shipment_detail', pk=pk)


class ShipmentReceiveView(LoginRequiredMixin, View):
    """Réceptionner une expédition"""
    def post(self, request, pk):
        shipment = get_object_or_404(Shipment, pk=pk)
        if shipment.status in [ShipmentStatus.SHIPPED, ShipmentStatus.IN_TRANSIT, ShipmentStatus.ARRIVED]:
            shipment.receive(request.user)
            messages.success(request, f'Expédition {shipment.reference} réceptionnée. Stock mis à jour.')
        return redirect('crm:shipment_detail', pk=pk)


class PreOrderListView(LoginRequiredMixin, ListView):
    """Liste des pré-commandes"""
    model = PreOrder
    template_name = 'crm/preorder_list.html'
    context_object_name = 'preorders'
    paginate_by = 20
    
    def get_queryset(self):
        qs = PreOrder.objects.select_related('customer', 'product', 'target_shipment')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = Product.objects.filter(is_active=True)
        return context


class PreOrderCreateView(LoginRequiredMixin, CreateView):
    """Créer une pré-commande"""
    model = PreOrder
    template_name = 'crm/preorder_form.html'
    fields = ['customer', 'product', 'quantity', 'target_shipment', 'notes']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['customers'] = Customer.objects.filter(is_active=True)
        context['products'] = Product.objects.filter(is_active=True)
        context['shipments'] = Shipment.objects.filter(
            status__in=[ShipmentStatus.PREPARING, ShipmentStatus.SHIPPED, ShipmentStatus.IN_TRANSIT]
        )
        return context
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Pré-commande créée')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('crm:preorder_list')


class PreOrderAllocateView(LoginRequiredMixin, View):
    """Allouer du stock aux pré-commandes par priorité"""
    def post(self, request):
        product_id = request.POST.get('product')
        
        if not product_id:
            messages.error(request, 'Sélectionnez un produit')
            return redirect('crm:preorder_list')
        
        product = get_object_or_404(Product, pk=product_id)
        available_stock = product.stock_quantity
        
        # Récupérer les pré-commandes en attente par priorité décroissante
        preorders = PreOrder.objects.filter(
            product=product,
            status=PreOrderStatus.PENDING
        ).order_by('-priority', 'created_at')
        
        allocated_count = 0
        for preorder in preorders:
            remaining = preorder.remaining_quantity
            if remaining > 0 and available_stock >= remaining:
                preorder.quantity_allocated = preorder.quantity
                preorder.status = PreOrderStatus.ALLOCATED
                preorder.save()
                
                # Créer réservation
                StockReservation.objects.create(
                    product=product,
                    customer=preorder.customer,
                    quantity=remaining,
                    reservation_type='preorder',
                    preorder=preorder,
                    created_by=request.user
                )
                
                available_stock -= remaining
                allocated_count += 1
            elif remaining > 0 and available_stock > 0:
                # Allocation partielle
                preorder.quantity_allocated += available_stock
                preorder.save()
                
                StockReservation.objects.create(
                    product=product,
                    customer=preorder.customer,
                    quantity=available_stock,
                    reservation_type='preorder',
                    preorder=preorder,
                    created_by=request.user
                )
                available_stock = 0
                break
        
        messages.success(request, f'{allocated_count} pré-commande(s) allouée(s)')
        return redirect('crm:preorder_list')


class InventoryListView(LoginRequiredMixin, ListView):
    """Liste des contrôles d'inventaire"""
    model = InventoryControl
    template_name = 'crm/inventory_list.html'
    context_object_name = 'inventories'
    paginate_by = 20


class InventoryCreateView(LoginRequiredMixin, View):
    """Créer un nouveau contrôle d'inventaire"""
    def get(self, request):
        return render(request, 'crm/inventory_form.html', {
            'products': Product.objects.filter(is_active=True)
        })
    
    def post(self, request):
        name = request.POST.get('name', 'Inventaire')
        control_date = request.POST.get('control_date', timezone.now().date())
        
        inventory = InventoryControl.objects.create(
            name=name,
            control_date=control_date,
            created_by=request.user
        )
        
        # Créer les lignes pour tous les produits actifs
        for product in Product.objects.filter(is_active=True):
            InventoryLine.objects.create(
                inventory=inventory,
                product=product,
                theoretical_quantity=product.stock_quantity
            )
        
        messages.success(request, f'Inventaire {inventory.reference} créé')
        return redirect('crm:inventory_detail', pk=inventory.pk)


class InventoryDetailView(LoginRequiredMixin, DetailView):
    model = InventoryControl
    template_name = 'crm/inventory_detail.html'
    context_object_name = 'inventory'
    
    def post(self, request, pk):
        """Mettre à jour les quantités physiques"""
        inventory = self.get_object()
        
        if inventory.status != InventoryControlStatus.DRAFT:
            messages.error(request, 'Cet inventaire ne peut plus être modifié')
            return redirect('crm:inventory_detail', pk=pk)
        
        # Mettre à jour les quantités physiques
        for line in inventory.lines.all():
            qty = request.POST.get(f'qty_{line.pk}')
            if qty:
                try:
                    line.physical_quantity = Decimal(qty)
                    line.save()
                except (ValueError, InvalidOperation):
                    pass
        
        inventory.status = InventoryControlStatus.IN_PROGRESS
        inventory.save()
        
        messages.success(request, 'Quantités enregistrées')
        return redirect('crm:inventory_detail', pk=pk)


class InventoryValidateView(LoginRequiredMixin, View):
    """Valider un inventaire et appliquer les ajustements"""
    def post(self, request, pk):
        inventory = get_object_or_404(InventoryControl, pk=pk)
        
        if inventory.status not in [InventoryControlStatus.DRAFT, InventoryControlStatus.IN_PROGRESS]:
            messages.error(request, 'Cet inventaire ne peut pas être validé')
            return redirect('crm:inventory_detail', pk=pk)
        
        inventory.validate(request.user)
        messages.success(request, f'Inventaire {inventory.reference} validé. Stock ajusté.')
        return redirect('crm:inventory_detail', pk=pk)


# ============================================================
# PHASE 5: LOGISTIQUE - BONS DE LIVRAISON
# ============================================================

from .models import DeliveryNote, DeliveryNoteItem, DeliveryNoteStatus


class DeliveryNoteListView(LoginRequiredMixin, ListView):
    """Liste des bons de livraison"""
    model = DeliveryNote
    template_name = 'crm/delivery_list.html'
    context_object_name = 'deliveries'
    paginate_by = 20
    
    def get_queryset(self):
        qs = DeliveryNote.objects.select_related('customer', 'order')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = DeliveryNoteStatus.choices
        return context


class DeliveryNoteCreateView(LoginRequiredMixin, View):
    """Créer un BL depuis une commande"""
    def get(self, request):
        # Liste des commandes sans BL complet
        orders = Order.objects.filter(
            status__in=[OrderStatus.CONFIRMED, OrderStatus.PROCESSING, OrderStatus.READY]
        ).exclude(
            delivery_notes__status=DeliveryNoteStatus.DELIVERED
        ).select_related('customer')
        return render(request, 'crm/delivery_form.html', {'orders': orders})
    
    def post(self, request):
        order_id = request.POST.get('order')
        if not order_id:
            messages.error(request, 'Sélectionnez une commande')
            return redirect('crm:delivery_create')
        
        order = get_object_or_404(Order, pk=order_id)
        
        # Créer le BL
        delivery = DeliveryNote.create_from_order(order, request.user)
        
        # Copier adresse personnalisée si fournie
        custom_address = request.POST.get('delivery_address')
        if custom_address:
            delivery.delivery_address = custom_address
            delivery.save()
        
        messages.success(request, f'Bon de livraison {delivery.number} créé')
        return redirect('crm:delivery_detail', pk=delivery.pk)


class DeliveryNoteDetailView(LoginRequiredMixin, DetailView):
    model = DeliveryNote
    template_name = 'crm/delivery_detail.html'
    context_object_name = 'delivery'


class DeliveryNoteReadyView(LoginRequiredMixin, View):
    """Marquer un BL prêt pour livraison"""
    def post(self, request, pk):
        delivery = get_object_or_404(DeliveryNote, pk=pk)
        if delivery.status == DeliveryNoteStatus.DRAFT:
            delivery.status = DeliveryNoteStatus.READY
            delivery.save()
            messages.success(request, f'BL {delivery.number} prêt pour livraison')
        return redirect('crm:delivery_detail', pk=pk)


class DeliveryNoteStartView(LoginRequiredMixin, View):
    """Démarrer la livraison"""
    def post(self, request, pk):
        delivery = get_object_or_404(DeliveryNote, pk=pk)
        if delivery.status in [DeliveryNoteStatus.DRAFT, DeliveryNoteStatus.READY]:
            delivery.status = DeliveryNoteStatus.IN_DELIVERY
            delivery.transporter = request.POST.get('transporter', '')
            delivery.vehicle_info = request.POST.get('vehicle_info', '')
            delivery.planned_date = timezone.now().date()
            delivery.save()
            messages.success(request, f'Livraison {delivery.number} démarrée')
        return redirect('crm:delivery_detail', pk=pk)


class DeliveryNoteDeliverView(LoginRequiredMixin, View):
    """Confirmer la livraison"""
    def post(self, request, pk):
        delivery = get_object_or_404(DeliveryNote, pk=pk)
        if delivery.status in [DeliveryNoteStatus.READY, DeliveryNoteStatus.IN_DELIVERY]:
            received_by = request.POST.get('received_by', '')
            delivery.mark_delivered(request.user, received_by)
            messages.success(request, f'BL {delivery.number} livré avec succès')
        return redirect('crm:delivery_detail', pk=pk)


class DeliveryNotePDFView(LoginRequiredMixin, View):
    """Générer le PDF du bon de livraison"""
    def get(self, request, pk):
        delivery = get_object_or_404(DeliveryNote, pk=pk)
        return render(request, 'crm/delivery_pdf.html', {
            'delivery': delivery,
            'print_mode': True
        })


# ============================================================
# PHASE 6: EMAILING & AUTOMATISATIONS
# ============================================================

from .models import EmailTemplate, EmailLog, EmailStatus, AutomationRule, trigger_automation


class EmailTemplateListView(LoginRequiredMixin, ListView):
    """Liste des templates email"""
    model = EmailTemplate
    template_name = 'crm/email_template_list.html'
    context_object_name = 'templates'


class EmailTemplateDetailView(LoginRequiredMixin, DetailView):
    """Détail d'un template avec prévisualisation"""
    model = EmailTemplate
    template_name = 'crm/email_template_detail.html'
    context_object_name = 'template'


class EmailLogListView(LoginRequiredMixin, ListView):
    """Historique des emails envoyés"""
    model = EmailLog
    template_name = 'crm/email_log_list.html'
    context_object_name = 'logs'
    paginate_by = 50
    
    def get_queryset(self):
        qs = EmailLog.objects.select_related('customer', 'template')
        customer = self.request.GET.get('customer')
        if customer:
            qs = qs.filter(customer_id=customer)
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = EmailStatus.choices
        return context


class EmailComposeView(LoginRequiredMixin, View):
    """Composer et envoyer un email"""
    def get(self, request):
        customers = Customer.objects.filter(is_active=True, email__isnull=False).exclude(email='')
        templates = EmailTemplate.objects.filter(is_active=True)
        return render(request, 'crm/email_compose.html', {
            'customers': customers,
            'templates': templates
        })
    
    def post(self, request):
        template_id = request.POST.get('template')
        customer_id = request.POST.get('customer')
        subject = request.POST.get('subject')
        body = request.POST.get('body')
        
        if not customer_id:
            messages.error(request, 'Sélectionnez un destinataire')
            return redirect('crm:email_compose')
        
        customer = get_object_or_404(Customer, pk=customer_id)
        if not customer.email:
            messages.error(request, 'Ce client n\'a pas d\'adresse email')
            return redirect('crm:email_compose')
        
        # Créer et envoyer l'email
        template = None
        if template_id:
            template = EmailTemplate.objects.filter(pk=template_id).first()
        
        log = EmailLog.objects.create(
            template=template,
            customer=customer,
            recipient_email=customer.email,
            recipient_name=customer.name,
            subject=subject,
            body_html=body,
            sent_by=request.user
        )
        
        if log.send():
            messages.success(request, f'Email envoyé à {customer.email}')
        else:
            messages.error(request, f'Erreur d\'envoi: {log.error_message}')
        
        return redirect('crm:email_logs')


class EmailPreviewView(LoginRequiredMixin, View):
    """Prévisualiser un template avec données de test"""
    def get(self, request, pk):
        template = get_object_or_404(EmailTemplate, pk=pk)
        
        # Données de test
        context = {
            'client_name': 'Client Test',
            'order_number': 'CMD-202501-0001',
            'order_date': '02/01/2025',
            'amount': '500 000',
            'items_count': 5,
            'invoice_number': 'SGS-202501-0001',
            'due_date': '15/01/2025',
            'payment_date': '10/01/2025',
            'payment_ref': 'PAY-20250110-0001',
            'delivery_date': '05/01/2025',
            'delivery_number': 'BL-202501-0001',
            'products_list': 'Farine de Poisson Premium (100 sacs)',
            'promo_title': 'Offre Spéciale Janvier',
            'promo_description': 'Profitez de prix exceptionnels !',
            'promo_discount': '-10%',
            'promo_end_date': '31/01/2025',
            'orders_count': 25,
            'loyalty_reward': '5% de remise sur votre prochaine commande',
            'order_link': '#'
        }
        
        subject, body_html, _ = template.render(context)
        
        return render(request, 'crm/email_preview.html', {
            'template': template,
            'rendered_subject': subject,
            'rendered_body': body_html
        })


class AutomationRuleListView(LoginRequiredMixin, ListView):
    """Liste des règles d'automatisation"""
    model = AutomationRule
    template_name = 'crm/automation_list.html'
    context_object_name = 'rules'


class AutomationRuleToggleView(LoginRequiredMixin, View):
    """Activer/désactiver une règle"""
    def post(self, request, pk):
        rule = get_object_or_404(AutomationRule, pk=pk)
        rule.is_active = not rule.is_active
        rule.save()
        status = 'activée' if rule.is_active else 'désactivée'
        messages.success(request, f'Règle {rule.name} {status}')
        return redirect('crm:automation_list')


# ============================================================
# PHASE 7: GOUVERNANCE & REPORTING
# ============================================================

from .models import (AuditLog, AuditAction, ApprovalRequest, ApprovalStatus,
                     UserProfile, UserRole, KPISnapshot)


class KPIDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard KPIs direction"""
    template_name = 'crm/kpi_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        # Capturer les KPIs du jour
        snapshot = KPISnapshot.capture(today)
        context['today'] = snapshot
        
        # Historique 7 derniers jours
        week_ago = today - timedelta(days=7)
        context['week_snapshots'] = KPISnapshot.objects.filter(date__gte=week_ago)
        
        # Statistiques temps réel
        context['total_customers'] = Customer.objects.filter(is_active=True).count()
        context['total_products'] = Product.objects.filter(is_active=True).count()
        
        # Commandes en cours
        context['pending_orders'] = Order.objects.filter(
            status__in=[OrderStatus.DRAFT, OrderStatus.CONFIRMED, OrderStatus.PROCESSING]
        ).count()
        
        # Top 5 clients par CA
        context['top_customers'] = Customer.objects.annotate(
            total_ca=Sum('orders__total')
        ).filter(total_ca__gt=0).order_by('-total_ca')[:5]
        
        # Produits les plus vendus
        context['top_products'] = Product.objects.annotate(
            total_sold=Sum('orderitem__quantity')
        ).filter(total_sold__gt=0).order_by('-total_sold')[:5]
        
        # Approbations en attente
        context['pending_approvals'] = ApprovalRequest.objects.filter(
            status=ApprovalStatus.PENDING
        ).count()
        
        return context


class AuditLogListView(LoginRequiredMixin, ListView):
    """Journal d'audit"""
    model = AuditLog
    template_name = 'crm/audit_log.html'
    context_object_name = 'logs'
    paginate_by = 100
    
    def get_queryset(self):
        qs = AuditLog.objects.select_related('user')
        action = self.request.GET.get('action')
        if action:
            qs = qs.filter(action=action)
        model = self.request.GET.get('model')
        if model:
            qs = qs.filter(model_name=model)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['action_choices'] = AuditAction.choices
        context['models'] = AuditLog.objects.values_list('model_name', flat=True).distinct()
        return context


class ApprovalListView(LoginRequiredMixin, ListView):
    """Liste des demandes d'approbation"""
    model = ApprovalRequest
    template_name = 'crm/approval_list.html'
    context_object_name = 'approvals'
    paginate_by = 20
    
    def get_queryset(self):
        qs = ApprovalRequest.objects.select_related('requested_by', 'approved_by')
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        else:
            # Par défaut: en attente
            qs = qs.filter(status=ApprovalStatus.PENDING)
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = ApprovalStatus.choices
        return context


class ApprovalActionView(LoginRequiredMixin, View):
    """Approuver ou rejeter une demande"""
    def post(self, request, pk):
        approval = get_object_or_404(ApprovalRequest, pk=pk)
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        
        if action == 'approve':
            approval.approve(request.user, notes)
            messages.success(request, 'Demande approuvée')
        elif action == 'reject':
            approval.reject(request.user, notes)
            messages.warning(request, 'Demande rejetée')
        
        return redirect('crm:approval_list')


class UserProfileListView(LoginRequiredMixin, ListView):
    """Gestion des utilisateurs et rôles"""
    model = UserProfile
    template_name = 'crm/user_list.html'
    context_object_name = 'profiles'
    
    def get_queryset(self):
        # Créer profils manquants
        User = get_user_model()
        for user in User.objects.filter(profile__isnull=True):
            UserProfile.objects.create(user=user)
        return UserProfile.objects.select_related('user')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['role_choices'] = UserRole.choices
        return context


class UserRoleUpdateView(LoginRequiredMixin, View):
    """Modifier le rôle d'un utilisateur"""
    def post(self, request, pk):
        profile = get_object_or_404(UserProfile, pk=pk)
        role = request.POST.get('role')
        limit = request.POST.get('approval_limit', 0)
        
        if role in dict(UserRole.choices):
            profile.role = role
            try:
                profile.approval_limit = Decimal(limit)
            except:
                pass
            profile.save()
            
            # Log audit
            AuditLog.log(request.user, AuditAction.UPDATE, profile, {
                'role': role,
                'approval_limit': str(profile.approval_limit)
            })
            
            messages.success(request, f'Rôle de {profile.user.username} mis à jour')
        
        return redirect('crm:user_list')


class ReportView(LoginRequiredMixin, TemplateView):
    """Rapports et exports"""
    template_name = 'crm/reports.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Période
        period = self.request.GET.get('period', 'month')
        today = timezone.now().date()
        
        if period == 'week':
            start_date = today - timedelta(days=7)
        elif period == 'month':
            start_date = today - timedelta(days=30)
        elif period == 'quarter':
            start_date = today - timedelta(days=90)
        elif period == 'year':
            start_date = today - timedelta(days=365)
        else:
            start_date = today - timedelta(days=30)
        
        context['period'] = period
        context['start_date'] = start_date
        context['end_date'] = today
        
        # Ventes
        orders = Order.objects.filter(created_at__date__gte=start_date)
        context['orders_count'] = orders.count()
        context['orders_total'] = orders.aggregate(total=Sum('total'))['total'] or 0
        
        # Par statut
        context['orders_by_status'] = orders.values('status').annotate(
            count=Count('id'),
            total=Sum('total')
        )
        
        # Paiements
        payments = Payment.objects.filter(payment_date__gte=start_date)
        context['payments_total'] = payments.aggregate(total=Sum('amount'))['total'] or 0
        context['payments_by_method'] = payments.values('payment_method').annotate(
            count=Count('id'),
            total=Sum('amount')
        )
        
        # Clients
        context['new_customers'] = Customer.objects.filter(
            created_at__date__gte=start_date
        ).count()
        
        # Livraisons
        context['deliveries_count'] = DeliveryNote.objects.filter(
            delivery_date__date__gte=start_date,
            status=DeliveryNoteStatus.DELIVERED
        ).count()
        
        return context


# Import models for F expression
from django.db import models
from django.urls import reverse


# ==============================================================
# PHASE 8: Optimisation UX & Productivité
# ==============================================================

class GlobalSearchView(LoginRequiredMixin, View):
    """Recherche globale - accessible via Ctrl+K"""
    
    def get(self, request):
        query = request.GET.get('q', '').strip()
        
        if len(query) < 2:
            return JsonResponse({'results': []})
        
        results = []
        
        # Recherche clients
        customers = Customer.objects.filter(
            Q(name__icontains=query) |
            Q(email__icontains=query) |
            Q(phone__icontains=query) |
            Q(company__icontains=query)
        )[:5]
        
        for c in customers:
            results.append({
                'type': 'customer',
                'icon': 'bi-person',
                'label': c.name,
                'sublabel': f'{c.customer_type} - {c.phone or c.email or ""}',
                'url': reverse('crm:customer_detail', args=[c.pk])
            })
        
        # Recherche commandes
        orders = Order.objects.filter(
            Q(order_number__icontains=query) |
            Q(customer__name__icontains=query)
        ).select_related('customer')[:5]
        
        for o in orders:
            results.append({
                'type': 'order',
                'icon': 'bi-cart',
                'label': f'Commande {o.order_number}',
                'sublabel': f'{o.customer.name} - {o.total} FCFA',
                'url': reverse('crm:order_detail', args=[o.pk])
            })
        
        # Recherche produits
        products = Product.objects.filter(
            Q(name__icontains=query) |
            Q(sku__icontains=query)
        )[:5]
        
        for p in products:
            results.append({
                'type': 'product',
                'icon': 'bi-box',
                'label': p.name,
                'sublabel': f'{p.stock_quantity} en stock - {p.price_b2b} FCFA',
                'url': reverse('crm:product_detail', args=[p.pk])
            })
        
        # Recherche factures
        invoices = Invoice.objects.filter(
            Q(invoice_number__icontains=query) |
            Q(customer__name__icontains=query)
        ).select_related('customer')[:5]
        
        for inv in invoices:
            results.append({
                'type': 'invoice',
                'icon': 'bi-receipt',
                'label': f'Facture {inv.invoice_number}',
                'sublabel': f'{inv.customer.name} - {inv.total} FCFA',
                'url': reverse('crm:invoice_detail', args=[inv.pk])
            })
        
        return JsonResponse({'results': results[:15]})


class QuickOrderCreateView(LoginRequiredMixin, View):
    """Création rapide de commande depuis fiche client"""
    
    def post(self, request, customer_pk):
        customer = get_object_or_404(Customer, pk=customer_pk)
        
        # Créer commande brouillon
        order = Order.objects.create(
            customer=customer,
            status=OrderStatus.DRAFT,
            payment_terms=customer.payment_terms,
            created_by=request.user
        )
        
        messages.success(request, f'Commande {order.order_number} créée pour {customer.name}')
        return redirect('crm:order_detail', pk=order.pk)


class QuickPaymentView(LoginRequiredMixin, View):
    """Encaissement rapide depuis commande/facture"""
    
    def post(self, request, invoice_pk):
        invoice = get_object_or_404(Invoice, pk=invoice_pk)
        
        amount = request.POST.get('amount')
        method = request.POST.get('method', 'cash')
        
        try:
            amount = Decimal(amount)
        except:
            messages.error(request, 'Montant invalide')
            return redirect('crm:invoice_detail', pk=invoice_pk)
        
        if amount <= 0:
            messages.error(request, 'Le montant doit être positif')
            return redirect('crm:invoice_detail', pk=invoice_pk)
        
        # Créer le paiement
        payment = Payment.objects.create(
            customer=invoice.customer,
            invoice=invoice,
            amount=amount,
            method=method,
            reference=f'ENC-{timezone.now().strftime("%Y%m%d%H%M%S")}',
            received_by=request.user,
            notes=f'Encaissement rapide facture {invoice.number}'
        )
        
        # Mettre à jour la facture
        invoice.amount_paid += amount
        if invoice.amount_paid >= invoice.total:
            invoice.status = InvoiceStatus.PAID
        elif invoice.amount_paid > 0:
            invoice.status = InvoiceStatus.PARTIAL
        invoice.save()
        
        # Créer entrée caisse
        CashRegisterEntry.objects.create(
            entry_type='in',
            amount=amount,
            payment_method=method,
            description=f'Paiement facture {invoice.number}',
            reference=payment.reference,
            payment=payment,
            created_by=request.user
        )
        
        messages.success(request, f'Paiement de {amount} FCFA enregistré')
        return redirect('crm:invoice_detail', pk=invoice_pk)


class CustomerStatsView(LoginRequiredMixin, TemplateView):
    """Statistiques détaillées d'un client"""
    template_name = 'crm/customer_stats.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = get_object_or_404(Customer, pk=self.kwargs['pk'])
        context['customer'] = customer
        
        # Commandes
        orders = Order.objects.filter(customer=customer)
        context['orders_count'] = orders.count()
        context['orders_total'] = orders.aggregate(total=Sum('total'))['total'] or 0
        
        # Par statut
        context['orders_by_status'] = orders.values('status').annotate(count=Count('id'))
        
        # Paiements
        payments = Payment.objects.filter(customer=customer)
        context['payments_total'] = payments.aggregate(total=Sum('amount'))['total'] or 0
        context['payments_count'] = payments.count()
        
        # Factures
        invoices = Invoice.objects.filter(customer=customer)
        context['invoices_count'] = invoices.count()
        context['invoices_unpaid'] = invoices.filter(
            status__in=[InvoiceStatus.SENT, InvoiceStatus.PARTIAL, InvoiceStatus.OVERDUE]
        ).aggregate(total=Sum(F('total') - F('amount_paid')))['total'] or 0
        
        # Historique mensuel (12 derniers mois)
        from django.db.models.functions import TruncMonth
        monthly_orders = orders.filter(
            created_at__gte=timezone.now() - timedelta(days=365)
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=Sum('total'),
            count=Count('id')
        ).order_by('month')
        context['monthly_orders'] = list(monthly_orders)
        
        # Avoirs
        avoirs = CustomerCredit.objects.filter(customer=customer, is_used=False)
        context['avoirs_total'] = avoirs.aggregate(total=Sum('amount'))['total'] or 0
        
        # Score client
        context['score'] = customer.calculate_score() if hasattr(customer, 'calculate_score') else None
        
        # Dernières commandes
        context['recent_orders'] = orders.order_by('-created_at')[:10]
        
        # Derniers paiements
        context['recent_payments'] = payments.order_by('-payment_date')[:10]
        
        return context




# ==============================================================
# PHASE 9: API & Exports
# ==============================================================

import csv
from django.http import HttpResponse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExportCustomersView(LoginRequiredMixin, View):
    """Export liste clients en Excel/CSV"""
    
    def get(self, request):
        format_type = request.GET.get('format', 'excel')
        customers = Customer.objects.filter(is_active=True).order_by('name')
        
        if format_type == 'csv':
            return self._export_csv(customers)
        else:
            return self._export_excel(customers)
    
    def _export_csv(self, customers):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="clients_serviac_{timezone.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')  # BOM UTF-8
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Nom', 'Type', 'Email', 'Téléphone', 'Ville', 'Limite Crédit', 'Solde', 'Score'])
        
        for c in customers:
            writer.writerow([
                c.name, c.customer_type, c.email or '', c.phone or '',
                c.city or '', c.credit_limit, c.balance, c.manual_score or ''
            ])
        
        return response
    
    def _export_excel(self, customers):
        if not OPENPYXL_AVAILABLE:
            return self._export_csv(customers)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients SERVIAC"
        
        # Style en-tête
        header_fill = PatternFill(start_color="1a5f2a", end_color="1a5f2a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['Nom', 'Type', 'Email', 'Téléphone', 'Ville', 'Limite Crédit', 'Solde', 'Score']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Données
        for row, c in enumerate(customers, 2):
            ws.cell(row=row, column=1, value=c.name)
            ws.cell(row=row, column=2, value=c.customer_type)
            ws.cell(row=row, column=3, value=c.email or '')
            ws.cell(row=row, column=4, value=c.phone or '')
            ws.cell(row=row, column=5, value=c.city or '')
            ws.cell(row=row, column=6, value=float(c.credit_limit))
            ws.cell(row=row, column=7, value=float(c.balance))
            ws.cell(row=row, column=8, value=c.manual_score or '')
        
        # Largeur colonnes
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 8
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="clients_serviac_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


class ExportOrdersView(LoginRequiredMixin, View):
    """Export commandes en Excel/CSV"""
    
    def get(self, request):
        format_type = request.GET.get('format', 'excel')
        
        # Filtres
        start_date = request.GET.get('start')
        end_date = request.GET.get('end')
        status = request.GET.get('status')
        
        orders = Order.objects.select_related('customer').order_by('-created_at')
        
        if start_date:
            orders = orders.filter(created_at__date__gte=start_date)
        if end_date:
            orders = orders.filter(created_at__date__lte=end_date)
        if status:
            orders = orders.filter(status=status)
        
        if format_type == 'csv':
            return self._export_csv(orders)
        else:
            return self._export_excel(orders)
    
    def _export_csv(self, orders):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="commandes_serviac_{timezone.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['N° Commande', 'Date', 'Client', 'Statut', 'Total HT', 'Total TTC', 'Payé'])
        
        for o in orders:
            writer.writerow([
                o.order_number, o.created_at.strftime('%d/%m/%Y'),
                o.customer.name, o.get_status_display(),
                o.subtotal, o.total, o.amount_paid
            ])
        
        return response
    
    def _export_excel(self, orders):
        if not OPENPYXL_AVAILABLE:
            return self._export_csv(orders)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Commandes SERVIAC"
        
        header_fill = PatternFill(start_color="1a5f2a", end_color="1a5f2a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['N° Commande', 'Date', 'Client', 'Statut', 'Total HT', 'Total TTC', 'Payé']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, o in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=o.order_number)
            ws.cell(row=row, column=2, value=o.created_at.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=o.customer.name)
            ws.cell(row=row, column=4, value=o.get_status_display())
            ws.cell(row=row, column=5, value=float(o.subtotal))
            ws.cell(row=row, column=6, value=float(o.total))
            ws.cell(row=row, column=7, value=float(o.amount_paid))
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="commandes_serviac_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


class ExportInvoicesView(LoginRequiredMixin, View):
    """Export factures en Excel/CSV"""
    
    def get(self, request):
        format_type = request.GET.get('format', 'excel')
        invoices = Invoice.objects.select_related('customer').order_by('-invoice_date')
        
        if format_type == 'csv':
            return self._export_csv(invoices)
        else:
            return self._export_excel(invoices)
    
    def _export_csv(self, invoices):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="factures_serviac_{timezone.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['N° Facture', 'Date', 'Échéance', 'Client', 'Statut', 'Total', 'Payé', 'Reste'])
        
        for inv in invoices:
            writer.writerow([
                inv.invoice_number, inv.invoice_date.strftime('%d/%m/%Y'),
                inv.due_date.strftime('%d/%m/%Y') if inv.due_date else '',
                inv.customer.name, inv.get_status_display(),
                inv.total, inv.amount_paid, inv.total - inv.amount_paid
            ])
        
        return response
    
    def _export_excel(self, invoices):
        if not OPENPYXL_AVAILABLE:
            return self._export_csv(invoices)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Factures SERVIAC"
        
        header_fill = PatternFill(start_color="1a5f2a", end_color="1a5f2a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['N° Facture', 'Date', 'Échéance', 'Client', 'Statut', 'Total', 'Payé', 'Reste']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, inv in enumerate(invoices, 2):
            ws.cell(row=row, column=1, value=inv.invoice_number)
            ws.cell(row=row, column=2, value=inv.invoice_date.strftime('%d/%m/%Y'))
            ws.cell(row=row, column=3, value=inv.due_date.strftime('%d/%m/%Y') if inv.due_date else '')
            ws.cell(row=row, column=4, value=inv.customer.name)
            ws.cell(row=row, column=5, value=inv.get_status_display())
            ws.cell(row=row, column=6, value=float(inv.total))
            ws.cell(row=row, column=7, value=float(inv.amount_paid))
            ws.cell(row=row, column=8, value=float(inv.total - inv.amount_paid))
        
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="factures_serviac_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


class ExportStockView(LoginRequiredMixin, View):
    """Export état du stock en Excel"""
    
    def get(self, request):
        format_type = request.GET.get('format', 'excel')
        products = Product.objects.filter(is_active=True).order_by('category__name', 'name')
        
        if format_type == 'csv':
            return self._export_csv(products)
        else:
            return self._export_excel(products)
    
    def _export_csv(self, products):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="stock_serviac_{timezone.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Référence', 'Produit', 'Catégorie', 'Stock actuel', 'Seuil alerte', 'Prix B2B', 'Prix B2C'])
        
        for p in products:
            writer.writerow([
                p.code, p.name, p.category.name if p.category else '',
                p.stock_quantity, p.alert_threshold,
                p.price_b2b, p.price_b2c
            ])
        
        return response
    
    def _export_excel(self, products):
        if not OPENPYXL_AVAILABLE:
            return self._export_csv(products)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock SERVIAC"
        
        header_fill = PatternFill(start_color="1a5f2a", end_color="1a5f2a", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        alert_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        headers = ['Référence', 'Produit', 'Catégorie', 'Stock', 'Seuil', 'Prix B2B', 'Prix B2C']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, p in enumerate(products, 2):
            ws.cell(row=row, column=1, value=p.code)
            ws.cell(row=row, column=2, value=p.name)
            ws.cell(row=row, column=3, value=p.category.name if p.category else '')
            ws.cell(row=row, column=4, value=float(p.stock_quantity))
            ws.cell(row=row, column=5, value=float(p.alert_threshold))
            ws.cell(row=row, column=6, value=float(p.price_b2b))
            ws.cell(row=row, column=7, value=float(p.price_b2c))
            
            # Surligner en rouge si stock < seuil
            if p.stock_quantity <= p.alert_threshold:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = alert_fill
        
        for col, width in enumerate([12, 35, 20, 10, 10, 12, 12], 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="stock_serviac_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        wb.save(response)
        return response


class ExportAgedBalanceView(LoginRequiredMixin, View):
    """Export balance âgée en Excel"""
    
    def get(self, request):
        from django.db.models import Case, When, Value, DecimalField
        
        customers = Customer.objects.filter(
            is_active=True, balance__gt=0
        ).order_by('-balance')
        
        wb = Workbook() if OPENPYXL_AVAILABLE else None
        
        if wb:
            ws = wb.active
            ws.title = "Balance Agée"
            
            header_fill = PatternFill(start_color="1a5f2a", end_color="1a5f2a", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            headers = ['Client', 'Total dû', '0-30j', '31-60j', '61-90j', '+90j']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            row = 2
            for c in customers:
                # Calcul par tranche (simplifié)
                ws.cell(row=row, column=1, value=c.name)
                ws.cell(row=row, column=2, value=float(c.balance))
                # Les colonnes par tranche nécessiteraient un calcul plus complexe
                ws.cell(row=row, column=3, value=0)
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=0)
                ws.cell(row=row, column=6, value=float(c.balance))  # Simplifié
                row += 1
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="balance_agee_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            wb.save(response)
            return response
        
        # Fallback CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="balance_agee_{timezone.now().strftime("%Y%m%d")}.csv"'
        response.write('\ufeff')
        
        writer = csv.writer(response, delimiter=';')
        writer.writerow(['Client', 'Total dû'])
        for c in customers:
            writer.writerow([c.name, c.balance])
        
        return response


# ============================================================
# VENTES EXPRESS (Phase 10)
# ============================================================

class SaleListView(LoginRequiredMixin, ListView):
    """Liste des ventes express"""
    template_name = 'crm/sale_list.html'
    context_object_name = 'sales'
    paginate_by = 20
    
    def get_queryset(self):
        from .models import Invoice, InvoiceStatus
        # Les ventes express sont les factures directes (sans commande préalable)
        qs = Invoice.objects.filter(
            order__isnull=True
        ).select_related('customer').order_by('-invoice_date')
        
        # Filtres
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        
        customer = self.request.GET.get('customer')
        if customer:
            qs = qs.filter(customer_id=customer)
        
        date_from = self.request.GET.get('date_from')
        if date_from:
            qs = qs.filter(date__gte=date_from)
        
        date_to = self.request.GET.get('date_to')
        if date_to:
            qs = qs.filter(date__lte=date_to)
        
        return qs
    
    def get_context_data(self, **kwargs):
        from .models import InvoiceStatus
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Ventes Express'
        context['status_choices'] = InvoiceStatus.choices
        context['customers'] = Customer.objects.filter(is_active=True).order_by('name')
        
        # Stats du jour
        today = timezone.now().date()
        today_sales = self.get_queryset().filter(invoice_date=today)
        context['today_count'] = today_sales.count()
        context['today_total'] = today_sales.aggregate(total=Sum('total'))['total'] or 0
        
        return context


class SaleCreateView(LoginRequiredMixin, View):
    """Création d'une vente express - Interface dynamique"""
    template_name = 'crm/sale_form.html'
    
    def get(self, request):
        from .models import CashRegister, CashRegisterStatus
        
        context = {
            'page_title': 'Nouvelle Vente Express',
            'customers': Customer.objects.filter(is_active=True, is_blocked=False).order_by('name'),
            'products': Product.objects.filter(is_active=True, stock_quantity__gt=0).select_related('category').order_by('name'),
        }
        
        # Vérifier si la caisse est ouverte
        open_register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
        context['cash_register_open'] = open_register is not None
        
        return render(request, self.template_name, context)
    
    def post(self, request):
        from .models import (
            Invoice, InvoiceItem, InvoiceStatus, Payment, PaymentMethod, 
            PaymentType, PaymentStatus, CashRegister, CashRegisterStatus,
            CashMovement, StockMovement, StockMovementType, CustomerLedger, AuditLog, AuditAction
        )
        import json
        
        try:
            # Récupérer les données
            customer_id = request.POST.get('customer_id')
            items_json = request.POST.get('items', '[]')
            payment_method = request.POST.get('payment_method', PaymentMethod.CASH)
            payment_reference = request.POST.get('payment_reference', '')
            notes = request.POST.get('notes', '')
            
            items = json.loads(items_json)
            
            if not items:
                messages.error(request, "Veuillez ajouter au moins un article.")
                return redirect('crm:sale_create')
            
            # Client (optionnel pour vente comptoir)
            customer = None
            if customer_id:
                customer = get_object_or_404(Customer, pk=customer_id)
            else:
                # Créer ou récupérer un client "Vente Comptoir" pour les ventes anonymes
                customer, created = Customer.objects.get_or_create(
                    name="Client Comptoir",
                    defaults={
                        'customer_type': CustomerType.B2C,
                        'phone': '0000000000',
                        'city': 'Abidjan',
                        'notes': 'Client générique pour les ventes comptoir anonymes'
                    }
                )
            
            # Créer la facture
            invoice = Invoice.objects.create(
                customer=customer,
                status=InvoiceStatus.PAID,
                invoice_date=timezone.now().date(),
                due_date=timezone.now().date(),
                notes=notes,
                created_by=request.user
            )
            
            total_ht = Decimal('0')
            
            # Créer les lignes de facture et déduire le stock
            for item in items:
                product = get_object_or_404(Product, pk=item['product_id'])
                quantity = Decimal(str(item['quantity']))
                unit_price = Decimal(str(item['unit_price']))
                
                # Vérifier le stock
                if product.stock_quantity < quantity:
                    invoice.delete()
                    messages.error(request, f"Stock insuffisant pour {product.name}")
                    return redirect('crm:sale_create')
                
                line_total = quantity * unit_price
                total_ht += line_total
                
                InvoiceItem.objects.create(
                    invoice=invoice,
                    product=product,
                    description=product.name,
                    quantity=quantity,
                    unit_price=unit_price,
                    line_total=line_total
                )
                
                # Mouvement de stock (sortie)
                StockMovement.objects.create(
                    product=product,
                    movement_type=StockMovementType.OUT,
                    quantity=-quantity,
                    reason=f"Vente express #{invoice.number}",
                    created_by=request.user
                )
                
                # Mettre à jour le stock
                product.stock_quantity -= quantity
                product.save()
            
            # Mettre à jour les totaux
            invoice.subtotal = total_ht
            invoice.total = total_ht  # Pas de TVA pour l'instant
            invoice.amount_paid = total_ht
            invoice.save()
            
            # Créer le paiement
            payment = Payment.objects.create(
                invoice=invoice,
                customer=customer,
                amount=total_ht,
                payment_method=payment_method,
                payment_type=PaymentType.PAYMENT,
                transaction_ref=payment_reference or '',
                notes=f"Paiement vente express #{invoice.number}",
                created_by=request.user
            )
            
            # Enregistrer en caisse si paiement espèces
            if payment_method == PaymentMethod.CASH:
                open_register = CashRegister.objects.filter(status=CashRegisterStatus.OPEN).first()
                if open_register:
                    CashMovement.objects.create(
                        cash_register=open_register,
                        movement_type='in',
                        amount=total_ht,
                        reason=f"Vente express #{invoice.number}",
                        created_by=request.user
                    )
                    open_register.expected_balance += total_ht
                    open_register.total_cash += total_ht
                    open_register.save()
            
            # Audit log
            AuditLog.objects.create(
                user=request.user,
                user_name=request.user.get_full_name() or request.user.username,
                action=AuditAction.CREATE,
                model_name='Invoice',
                object_id=str(invoice.id),
                object_repr=str(invoice),
                changes={'type': 'Vente express', 'total': str(total_ht)}
            )
            
            messages.success(request, f"Vente #{invoice.number} créée avec succès! Total: {total_ht:,.0f} FCFA")
            return redirect('crm:sale_detail', pk=invoice.pk)
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création de la vente: {str(e)}")
            return redirect('crm:sale_create')


class SaleDetailView(LoginRequiredMixin, DetailView):
    """Détail d'une vente express"""
    template_name = 'crm/sale_detail.html'
    context_object_name = 'sale'
    
    def get_queryset(self):
        from .models import Invoice
        return Invoice.objects.filter(order__isnull=True).select_related('customer', 'created_by')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = f'Vente #{self.object.number}'
        context['items'] = self.object.items.select_related('product')
        context['payments'] = self.object.payments.all()
        return context


class SalePDFView(LoginRequiredMixin, View):
    """Génération PDF d'une vente express"""
    
    def get(self, request, pk):
        from .models import Invoice
        from django.http import HttpResponse
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        import io
        import os
        
        invoice = get_object_or_404(Invoice, pk=pk, order__isnull=True)
        
        # Créer le buffer
        buffer = io.BytesIO()
        
        # Créer le document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1.5*cm,
            leftMargin=1.5*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        # Styles
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=20,
            alignment=TA_CENTER,
            spaceAfter=20,
            textColor=colors.HexColor('#1e3a5f')
        )
        
        header_style = ParagraphStyle(
            'Header',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#666666')
        )
        
        normal_style = ParagraphStyle(
            'NormalCustom',
            parent=styles['Normal'],
            fontSize=10,
            leading=14
        )
        
        bold_style = ParagraphStyle(
            'Bold',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold'
        )
        
        # Contenu
        elements = []
        
        # Logo et en-tête avec logo image
        from django.conf import settings
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo-serviac.png')
        
        # En-tête avec logo
        header_data = []
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=3.5*cm, height=3.5*cm)
            header_data = [[logo, [
                Paragraph("SERVIAC GROUP SUARL", title_style),
                Paragraph("Farine de Poisson Premium - Nutrition Animale", header_style),
                Paragraph("N'Dotré, près de Hotel Dandy, Abidjan - Côte d'Ivoire", header_style),
                Paragraph("Tél: +225 07 79 05 71 01 / +225 07 01 80 80 49", header_style),
                Paragraph("Email: info@serviac-group.com", header_style),
            ]]]
            header_table = Table(header_data, colWidths=[4*cm, 14*cm])
            header_table.setStyle(TableStyle([
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (0,0), 'CENTER'),
            ]))
            elements.append(header_table)
        else:
            elements.append(Paragraph("SERVIAC GROUP SUARL", title_style))
            elements.append(Paragraph("Farine de Poisson Premium - Nutrition Animale", header_style))
            elements.append(Paragraph("N'Dotré, près de Hotel Dandy, Abidjan - Côte d'Ivoire", header_style))
            elements.append(Paragraph("Tél: +225 07 79 05 71 01 / +225 07 01 80 80 49", header_style))
            elements.append(Paragraph("Email: info@serviac-group.com", header_style))
        elements.append(Spacer(1, 0.3*cm))
        
        # Ligne de séparation
        elements.append(Table([['']], colWidths=[18*cm], style=[
            ('LINEBELOW', (0,0), (-1,-1), 2, colors.HexColor('#d4a84b'))
        ]))
        elements.append(Spacer(1, 0.5*cm))
        
        # Type de document
        elements.append(Paragraph(f"<b>FACTURE N° {invoice.number}</b>", ParagraphStyle(
            'DocTitle', parent=styles['Heading2'], fontSize=16, alignment=TA_CENTER,
            textColor=colors.HexColor('#1e3a5f')
        )))
        elements.append(Spacer(1, 0.3*cm))
        
        # Informations facture et client
        info_data = [
            [Paragraph(f"<b>Date:</b> {invoice.invoice_date.strftime('%d/%m/%Y')}", normal_style),
             Paragraph(f"<b>Client:</b> {invoice.customer.name if invoice.customer else 'Vente Comptoir'}", normal_style)],
            [Paragraph(f"<b>Échéance:</b> {invoice.due_date.strftime('%d/%m/%Y')}", normal_style),
             Paragraph(f"<b>Tél:</b> {invoice.customer.phone if invoice.customer else '-'}", normal_style)],
        ]
        
        if invoice.customer and invoice.customer.address:
            info_data.append([
                Paragraph("", normal_style),
                Paragraph(f"<b>Adresse:</b> {invoice.customer.address}", normal_style)
            ])
        
        info_table = Table(info_data, colWidths=[9*cm, 9*cm])
        info_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Tableau des articles
        items = invoice.items.select_related('product').all()
        
        # Style pour en-tête blanc
        header_white_style = ParagraphStyle('HeaderWhite', parent=bold_style, textColor=colors.white)
        
        table_data = [
            [Paragraph('<b>Réf.</b>', header_white_style),
             Paragraph('<b>Désignation</b>', header_white_style),
             Paragraph('<b>Qté</b>', header_white_style),
             Paragraph('<b>P.U. (FCFA)</b>', header_white_style),
             Paragraph('<b>Total (FCFA)</b>', header_white_style)]
        ]
        
        for item in items:
            table_data.append([
                Paragraph(item.product.code if item.product else '-', normal_style),
                Paragraph(item.description or (item.product.name if item.product else '-'), normal_style),
                Paragraph(f"{item.quantity:,.0f}", normal_style),
                Paragraph(f"{item.unit_price:,.0f}", normal_style),
                Paragraph(f"{item.line_total:,.0f}", normal_style)
            ])
        
        item_table = Table(table_data, colWidths=[2.5*cm, 8*cm, 2*cm, 2.75*cm, 2.75*cm])
        item_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 10),
            ('TOPPADDING', (0,0), (-1,0), 10),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        elements.append(item_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Totaux - alignement corrigé avec montant et FCFA sur même ligne
        total_right_style = ParagraphStyle('TotalRight', parent=bold_style, fontSize=10, alignment=TA_RIGHT)
        total_ttc_style = ParagraphStyle('TotalTTC', parent=bold_style, fontSize=11, textColor=colors.HexColor('#1e3a5f'), alignment=TA_RIGHT)
        
        totals_data = [
            ['', '', '', Paragraph('<b>Total HT:</b>', total_right_style), Paragraph(f"<b>{invoice.subtotal:,.0f} FCFA</b>", total_right_style)],
            ['', '', '', Paragraph('<b>TVA:</b>', total_right_style), Paragraph(f"<b>{invoice.tax_amount:,.0f} FCFA</b>", total_right_style)],
            ['', '', '', Paragraph('<b>TOTAL TTC:</b>', total_ttc_style), Paragraph(f"<b>{invoice.total:,.0f} FCFA</b>", total_ttc_style)],
        ]
        
        totals_table = Table(totals_data, colWidths=[2.5*cm, 8*cm, 2*cm, 2.75*cm, 2.75*cm])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LINEABOVE', (3,2), (-1,2), 2, colors.HexColor('#d4a84b')),
            ('TOPPADDING', (0,2), (-1,2), 8),
        ]))
        elements.append(totals_table)
        elements.append(Spacer(1, 1*cm))
        
        # Statut paiement
        if invoice.amount_paid >= invoice.total:
            status_text = "✓ PAYÉE"
            status_color = colors.HexColor('#28a745')
        else:
            status_text = f"Reste à payer: {(invoice.total - invoice.amount_paid):,.0f} FCFA"
            status_color = colors.HexColor('#dc3545')
        
        elements.append(Paragraph(f"<b>{status_text}</b>", ParagraphStyle(
            'Status', parent=styles['Normal'], fontSize=14, alignment=TA_CENTER, textColor=status_color
        )))
        elements.append(Spacer(1, 1*cm))
        
        # Notes
        if invoice.notes:
            elements.append(Paragraph(f"<b>Notes:</b> {invoice.notes}", normal_style))
            elements.append(Spacer(1, 0.5*cm))
        
        # Pied de page avec infos légales
        elements.append(Table([['']], colWidths=[18*cm], style=[
            ('LINEBELOW', (0,0), (-1,-1), 1, colors.HexColor('#cccccc'))
        ]))
        elements.append(Spacer(1, 0.3*cm))
        
        footer_style = ParagraphStyle('FooterInfo', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#666666'))
        
        elements.append(Paragraph("Merci pour votre confiance!", ParagraphStyle(
            'FooterThanks', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, textColor=colors.HexColor('#1e3a5f'), fontName='Helvetica-Bold'
        )))
        elements.append(Spacer(1, 0.2*cm))
        elements.append(Paragraph("SERVIAC GROUP SUARL | Tél: (+225) 07 79 05 71 01 / 07 01 80 80 49 | Email: info@serviac-group.com", footer_style))
        elements.append(Paragraph("Site Web: www.serviac-group.com | Situé à Carrefour Dandy, N'Dotré (Anyama) | Abidjan – CÔTE D'IVOIRE", footer_style))
        elements.append(Spacer(1, 0.15*cm))
        elements.append(Paragraph("<b>RCCM: CI-GRDBSM-2020-B-1011 | Compte Contribuable: 2018961 U</b>", ParagraphStyle(
            'FooterLegal', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.HexColor('#1e3a5f')
        )))
        
        # Générer le PDF
        doc.build(elements)
        
        # Retourner la réponse
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="facture_{invoice.number}.pdf"'
        
        return response


class SaleQuickSearchView(LoginRequiredMixin, View):
    """Recherche rapide de produits/clients pour vente express (AJAX)"""
    
    def get(self, request):
        query = request.GET.get('q', '').strip()
        search_type = request.GET.get('type', 'product')
        
        if len(query) < 2:
            return JsonResponse({'results': []})
        
        if search_type == 'product':
            products = Product.objects.filter(
                Q(name__icontains=query) | Q(code__icontains=query),
                is_active=True,
                stock_quantity__gt=0
            )[:10]
            
            results = [{
                'id': p.id,
                'code': p.code,
                'name': p.name,
                'price_b2b': float(p.price_b2b),
                'price_b2c': float(p.price_b2c),
                'stock': float(p.stock_quantity),
                'unit': p.get_unit_display()
            } for p in products]
        else:
            customers = Customer.objects.filter(
                Q(name__icontains=query) | Q(phone__icontains=query),
                is_active=True,
                is_blocked=False
            )[:10]
            
            results = [{
                'id': c.id,
                'name': c.name,
                'phone': c.phone,
                'type': c.customer_type,
                'balance': float(c.balance)
            } for c in customers]
        
        return JsonResponse({'results': results})
